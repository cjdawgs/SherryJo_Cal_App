from openpyxl import Workbook, load_workbook

from scripts.generate_closure_workbook import HEADERS, ROWS, build


def test_closure_exceptions_is_final_formula_free_worksheet(tmp_path):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "closure.xlsx"
    workbook = Workbook()
    workbook.active.title = "Existing Evidence"
    workbook.save(template)

    build(template, output)

    generated = load_workbook(output, read_only=True, data_only=False, keep_links=False)
    sheet = generated.worksheets[-1]
    assert generated.sheetnames == ["Existing Evidence", "Closure Exceptions"]
    assert sheet.max_row == 4 + len(ROWS)
    assert sheet.max_column == len(HEADERS)
    assert sheet["A2"].value.startswith("Gate 158 status: BLOCKED")
    assert not any(
        isinstance(cell.value, str) and cell.value.startswith("=")
        for row in sheet.iter_rows()
        for cell in row
    )