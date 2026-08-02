"""Generate a closure-exceptions Excel workbook.

Usage (standalone)::

    python scripts/generate_closure_workbook.py [template.xlsx] [output.xlsx]

When called as a library, use :func:`build`.
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Public constants – consumed by tests and the CLI entry-point
# ---------------------------------------------------------------------------

HEADERS: list[str] = [
    "Exception ID",
    "Description",
    "Category",
    "Status",
    "Owner",
    "Target Date",
    "Notes",
]

ROWS: list[list[str]] = [
    [
        "CE-001",
        "Missing signature on gate review form",
        "Process",
        "Open",
        "CJ Dawgs",
        "2026-09-01",
        "Awaiting sign-off",
    ],
    [
        "CE-002",
        "Dependency not formally closed in tracker",
        "Dependency",
        "In Review",
        "SherryJo",
        "2026-09-15",
        "Tracked in issue tracker",
    ],
    [
        "CE-003",
        "Outstanding action item from Sprint 12 retro",
        "Action Item",
        "Open",
        "Team Lead",
        "2026-09-30",
        "Escalated to PM",
    ],
]


# ---------------------------------------------------------------------------
# Core builder
# ---------------------------------------------------------------------------


def build(template_path: str | Path, output_path: str | Path) -> None:
    """Append a *Closure Exceptions* sheet to *template_path* and save to *output_path*.

    Sheet layout
    ~~~~~~~~~~~~
    Row 1  – report title
    Row 2  – gate status banner (starts with "Gate 158 status: BLOCKED")
    Row 3  – instructions for reviewers
    Row 4  – column headers (:data:`HEADERS`)
    Row 5+ – data rows (:data:`ROWS`)
    """
    wb = load_workbook(template_path)
    ws = wb.create_sheet("Closure Exceptions")

    # Row 1 – title
    ws["A1"] = "Closure Exceptions Report"

    # Row 2 – gate status banner
    ws["A2"] = "Gate 158 status: BLOCKED — operator sign-off required before release"

    # Row 3 – reviewer instructions
    ws["A3"] = "All exceptions below must be resolved or formally accepted before closure."

    # Row 4 – column headers
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=4, column=col_idx, value=header)

    # Rows 5+ – data rows
    for row_offset, row_data in enumerate(ROWS):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=5 + row_offset, column=col_idx, value=value)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    args = sys.argv[1:]
    if len(args) != 2:
        print("Usage: python scripts/generate_closure_workbook.py <template.xlsx> <output.xlsx>")
        sys.exit(1)
    build(args[0], args[1])
    print(f"Written: {args[1]}")
